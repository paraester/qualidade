import os
import pandas as pd
import tkinter as tk
from tkinter import messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import subprocess
import logging

# Ativar logs
logging.basicConfig(level=logging.INFO)

# Função para verificar se a pasta 'Dados' e o arquivo CSV existem
def verificar_arquivo_csv():
    pasta_dados = os.path.join(os.getcwd(), "Dados")
    if not os.path.exists(pasta_dados):
        return None
    for arquivo in os.listdir(pasta_dados):
        if arquivo.endswith('.csv'):
            return os.path.join(pasta_dados, arquivo)
    return None

# Função para gerar o arquivo de qualidade (executa o script login_e_exportar.py)
def gerar_arquivo_qualidade():
    try:
        logging.info("Executando o script login_e_exportar.py")
        script_dir = os.path.dirname(os.path.abspath(__file__))
        subprocess.run(['python3', 'login_e_exportar.py'], check=True, cwd=script_dir)
        messagebox.showinfo("Sucesso", "Arquivo de qualidade gerado com sucesso.")
    except subprocess.CalledProcessError as e:
        logging.error(f"Erro ao gerar o arquivo: {e}")
        messagebox.showerror("Erro", "Ocorreu um erro ao gerar o arquivo de qualidade.")

# Função para tratar o arquivo de qualidade (executa o script tratar.py)
def tratar_arquivo_qualidade():
    if verificar_arquivo_csv():
        try:
            logging.info("Executando o script tratar.py")
            script_dir = os.path.dirname(os.path.abspath(__file__))
            subprocess.run(['python3', 'tratar.py'], check=True, cwd=script_dir)
            messagebox.showinfo("Sucesso", "Arquivo tratado com sucesso.")
        except subprocess.CalledProcessError as e:
            logging.error(f"Erro ao tratar o arquivo: {e}")
            messagebox.showerror("Erro", "Ocorreu um erro ao tratar o arquivo.")
    else:
        messagebox.showwarning("Aviso", "Não foi encontrado o arquivo CSV. Gere o arquivo de qualidade primeiro.")

# Função para ler o arquivo CSV e identificar áreas responsáveis únicas
def ler_e_filtrar_areas_responsaveis():
    caminho_arquivo = verificar_arquivo_csv()
    if not caminho_arquivo:
        messagebox.showwarning("Aviso", "Não foi encontrado o arquivo CSV. Gere o arquivo de qualidade primeiro.")
        return None, None

    try:
        df = pd.read_csv(caminho_arquivo)
        if "Área Responsável" not in df.columns:
            messagebox.showerror("Erro", "A coluna 'Área Responsável' não foi encontrada no arquivo CSV.")
            return None, None

        areas_responsaveis = df["Área Responsável"].dropna().unique()
        logging.info(f"Ao todo foram encontradas {len(areas_responsaveis)} áreas responsáveis únicas.")
        return areas_responsaveis, df
    except Exception as e:
        logging.error(f"Erro ao ler o arquivo CSV: {e}")
        messagebox.showerror("Erro", "Ocorreu um erro ao ler o arquivo CSV.")
        return None, None

# Função para criar um novo arquivo CSV filtrado por Área Responsável
def criar_arquivo_filtrado(area_responsavel, df):
    try:
        df_filtrado = df[df["Área Responsável"] == area_responsavel]
        pasta_dados = os.path.join(os.getcwd(), "Dados")
        nome_arquivo = f"{area_responsavel.replace(' ', '_')}.csv"
        caminho_arquivo = os.path.join(pasta_dados, nome_arquivo)
        df_filtrado.to_csv(caminho_arquivo, index=False)
        messagebox.showinfo("Sucesso", f"Arquivo gerado para a área '{area_responsavel}' salvo em {caminho_arquivo}.")
    except Exception as e:
        logging.error(f"Erro ao criar o arquivo filtrado: {e}")
        messagebox.showerror("Erro", "Ocorreu um erro ao gerar o arquivo filtrado.")

# Função para inverter nome e sobrenome na coluna "Nome completo"
def inverter_nomes(df):
    if "Nome completo" not in df.columns:
        logging.error("Coluna 'Nome completo' não encontrada.")
        messagebox.showerror("Erro", "A coluna 'Nome completo' não foi encontrada no arquivo CSV.")
        return df

    df['Nome completo'] = df['Nome completo'].apply(lambda nome: inverter_nome_completo(nome))
    df = df.sort_values(by="Nome completo")
    return df

def inverter_nome_completo(nome):
    if pd.isna(nome):
        return nome
    partes = nome.split(',')
    if len(partes) == 2:
        return partes[1].strip() + ' ' + partes[0].strip()
    return nome

# Função para atualizar a combobox com as áreas responsáveis
def atualizar_combobox_areas(combobox):
    areas_responsaveis, df = ler_e_filtrar_areas_responsaveis()
    if areas_responsaveis is not None:
        combobox['values'] = list(areas_responsaveis)
    return df

# Função para selecionar a área e gerar o arquivo filtrado
def selecionar_area_e_gerar(combobox, df):
    area_selecionada = combobox.get()
    if area_selecionada and df is not None:
        criar_arquivo_filtrado(area_selecionada, df)
    else:
        messagebox.showwarning("Aviso", "Selecione uma área responsável antes de continuar.")

# Função para aplicar a inversão dos nomes e salvar o CSV modificado
def inverter_nomes_e_salvar():
    caminho_arquivo = verificar_arquivo_csv()
    if not caminho_arquivo:
        messagebox.showwarning("Aviso", "Não foi encontrado o arquivo CSV para inverter os nomes.")
        return

    try:
        df = pd.read_csv(caminho_arquivo)
        df_modificado = inverter_nomes(df)
        df_modificado.to_csv(caminho_arquivo, index=False)
        messagebox.showinfo("Sucesso", f"Os nomes foram invertidos e o arquivo foi salvo em {caminho_arquivo}.")
    except Exception as e:
        logging.error(f"Erro ao inverter os nomes e salvar o arquivo: {e}")
        messagebox.showerror("Erro", "Ocorreu um erro ao inverter os nomes e salvar o arquivo.")


from openpyxl.utils import get_column_letter

# Função para criar abas no Excel para cada pessoa com seus respectivos dados
def inverter_nomes_e_criar_abas_excel():
    caminho_arquivo = verificar_arquivo_csv()
    if not caminho_arquivo:
        messagebox.showwarning("Aviso", "Não foi encontrado o arquivo CSV.")
        return

    try:
        # Ler o arquivo CSV
        df = pd.read_csv(caminho_arquivo)
        df = inverter_nomes(df)  # Aplica a inversão e ordenação dos nomes, se necessário

        # Carregar o modelo da planilha existente
        modelo_path = 'modelo.xlsx'
        if not os.path.exists(modelo_path):
            messagebox.showerror("Erro", "O arquivo modelo.xlsx não foi encontrado.")
            return
        
        wb = load_workbook(modelo_path)
        if 'primeiroNome' not in wb.sheetnames:
            messagebox.showerror("Erro", "A aba 'primeiroNome' não foi encontrada no modelo.")
            return
        
        aba_modelo = wb['primeiroNome']
        abas_existentes = {}

        # Atualizar as colunas conforme solicitado
        colunas = ['Nome', 'Andamento', 'Início', 'Término', 'Início.1', 'Término.1', 'Alocação padrão']

        # Iterar sobre cada linha do DataFrame
        for _, row in df.iterrows():
            nome_completo = row['Nome completo']
            area_responsavel = row['Área Responsável']
            primeiro_nome = nome_completo.split()[0]  # Pegar o primeiro nome

            # Verificar se a aba já existe para o primeiro nome
            if primeiro_nome in abas_existentes:
                nova_aba = abas_existentes[primeiro_nome]
                # Encontra a primeira linha vazia na aba para continuar adicionando dados
                ultima_linha = len(list(nova_aba.iter_rows(min_row=9, max_row=nova_aba.max_row, values_only=True))) + 9
            else:
                # Criar uma nova aba a partir do modelo
                nova_aba = wb.copy_worksheet(aba_modelo)
                nova_aba.title = primeiro_nome
                abas_existentes[primeiro_nome] = nova_aba
                ultima_linha = 9  # Começar a inserir dados na linha 9

                # Substituir os placeholders na aba (Nome Completo e Área Responsável)
                for linha in nova_aba.iter_rows():
                    for cell in linha:
                        if isinstance(cell.value, str):
                            if "<Nome Completo>" in cell.value:
                                cell.value = cell.value.replace("<Nome Completo>", nome_completo)
                            if "<nome da área responsável>" in cell.value:
                                cell.value = cell.value.replace("<nome da área responsável>", area_responsavel)

            # Inserir os dados a partir da linha 9 em diante (sequencial)
            for i, coluna in enumerate(colunas, start=1):
                valor = row[coluna] if coluna in row else ""
                celula = nova_aba[f"{get_column_letter(i)}{ultima_linha}"]
                celula.value = valor

            # Incrementar `ultima_linha` após a inserção
            ultima_linha += 1

        # Salvar o arquivo Excel atualizado
        novo_arquivo_excel = os.path.join(os.getcwd(), "Dados", "resultado_planilha.xlsx")
        wb.save(novo_arquivo_excel)
        messagebox.showinfo("Sucesso", f"Planilha criada com abas para cada pessoa salva em {novo_arquivo_excel}")

    except Exception as e:
        logging.error(f"Erro ao criar as abas no Excel: {e}")
        messagebox.showerror("Erro", "Ocorreu um erro ao criar as abas no Excel.")






# Interface gráfica
def criar_interface():
    app = ttk.Window(themename='cosmo')
    app.title("Sistema de Qualidade")
    app.geometry("600x600")
    app.resizable(True, True)

    titulo = ttk.Label(app, text="Sistema de Qualidade", font=("Helvetica", 16), bootstyle=PRIMARY)
    titulo.pack(pady=10)

    btn_gerar = ttk.Button(app, text="Gerar arquivo Qualidade", command=gerar_arquivo_qualidade, bootstyle=SUCCESS)
    btn_gerar.pack(pady=10)

    btn_tratar = ttk.Button(app, text="Tratar arquivo Qualidade", command=tratar_arquivo_qualidade, bootstyle=INFO)
    btn_tratar.pack(pady=10)

    frame_combobox = ttk.Frame(app)
    frame_combobox.pack(pady=20)

    btn_carregar_areas = ttk.Button(frame_combobox, text="Carregar Áreas", bootstyle=PRIMARY)
    btn_carregar_areas.pack(pady=10)

    combobox_areas = ttk.Combobox(frame_combobox, bootstyle=INFO)
    combobox_areas.pack(pady=10)

    btn_gerar_filtrado = ttk.Button(frame_combobox, text="Gerar Arquivo Filtrado", bootstyle=SUCCESS)
    btn_gerar_filtrado.pack(pady=10)

    btn_inverter_nomes = ttk.Button(app, text="Inverter Nomes", command=inverter_nomes_e_salvar, bootstyle=WARNING)
    btn_inverter_nomes.pack(pady=10)

    btn_criar_abas = ttk.Button(app, text="Criar Abas no Excel", command=inverter_nomes_e_criar_abas_excel, bootstyle=WARNING)
    btn_criar_abas.pack(pady=10)

    def carregar_areas_e_df():
        df = atualizar_combobox_areas(combobox_areas)
        btn_gerar_filtrado.config(command=lambda: selecionar_area_e_gerar(combobox_areas, df))

    btn_carregar_areas.config(command=carregar_areas_e_df)

    app.mainloop()

if __name__ == "__main__":
    criar_interface()
